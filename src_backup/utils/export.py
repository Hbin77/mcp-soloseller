"""
내보내기 및 리포트 유틸리티
엑셀, CSV, PDF 생성
"""
import io
import csv
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


@dataclass
class ExportResult:
    """내보내기 결과"""
    success: bool
    filename: str
    content: bytes
    content_type: str
    error: Optional[str] = None


class ExcelExporter:
    """엑셀 내보내기"""
    
    @staticmethod
    def export_orders(orders: List[Dict], filename: str = None) -> ExportResult:
        """주문 목록 엑셀 내보내기"""
        if not EXCEL_AVAILABLE:
            return ExportResult(False, "", b"", "", "openpyxl 라이브러리가 설치되지 않았습니다")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "주문목록"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더
        headers = ["번호", "채널", "주문번호", "주문자", "수령인", "연락처", "주소", "금액", "상태", "송장번호", "주문일시"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value="스마트스토어" if order.get("channel") == "naver" else "쿠팡")
            ws.cell(row=row, column=3, value=order.get("channel_order_id", ""))
            ws.cell(row=row, column=4, value=order.get("buyer_name", ""))
            ws.cell(row=row, column=5, value=order.get("receiver_name", ""))
            ws.cell(row=row, column=6, value=order.get("receiver_phone", ""))
            ws.cell(row=row, column=7, value=order.get("receiver_address", ""))
            ws.cell(row=row, column=8, value=order.get("total_amount", 0))
            ws.cell(row=row, column=9, value=order.get("status", ""))
            ws.cell(row=row, column=10, value=order.get("tracking_number", ""))
            ws.cell(row=row, column=11, value=order.get("ordered_at", ""))
        
        # 열 너비 조정
        column_widths = [6, 12, 20, 10, 10, 15, 40, 12, 10, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        if not filename:
            filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    @staticmethod
    def export_products(products: List[Dict], filename: str = None) -> ExportResult:
        """상품 목록 엑셀 내보내기"""
        if not EXCEL_AVAILABLE:
            return ExportResult(False, "", b"", "", "openpyxl 라이브러리가 설치되지 않았습니다")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "상품목록"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # 헤더
        headers = ["번호", "SKU", "상품명", "재고", "임계값", "가격", "네이버ID", "쿠팡ID", "상태"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 데이터
        low_stock_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=product.get("sku", ""))
            ws.cell(row=row, column=3, value=product.get("name", ""))
            
            stock_cell = ws.cell(row=row, column=4, value=product.get("stock_quantity", 0))
            threshold = product.get("stock_alert_threshold", 5)
            if product.get("stock_quantity", 0) <= threshold:
                stock_cell.fill = low_stock_fill
            
            ws.cell(row=row, column=5, value=threshold)
            ws.cell(row=row, column=6, value=product.get("price", 0))
            ws.cell(row=row, column=7, value=product.get("naver_product_id", ""))
            ws.cell(row=row, column=8, value=product.get("coupang_product_id", ""))
            ws.cell(row=row, column=9, value="활성" if product.get("is_active", True) else "비활성")
        
        # 열 너비
        column_widths = [6, 15, 40, 8, 8, 12, 15, 15, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        if not filename:
            filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    @staticmethod
    def export_daily_report(report: Dict, filename: str = None) -> ExportResult:
        """일일 리포트 엑셀 내보내기"""
        if not EXCEL_AVAILABLE:
            return ExportResult(False, "", b"", "", "openpyxl 라이브러리가 설치되지 않았습니다")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "일일리포트"
        
        # 제목
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"📊 일일 리포트 - {report.get('date', '')}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # 주문 현황
        ws['A3'] = "📦 주문 현황"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "총 주문"
        ws['B4'] = report.get('orders', {}).get('total', 0)
        ws['A5'] = "스마트스토어"
        ws['B5'] = report.get('orders', {}).get('naver', 0)
        ws['A6'] = "쿠팡"
        ws['B6'] = report.get('orders', {}).get('coupang', 0)
        
        # 매출 현황
        ws['A8'] = "💰 매출 현황"
        ws['A8'].font = Font(bold=True, size=12)
        
        ws['A9'] = "총 매출"
        ws['B9'] = report.get('sales', {}).get('total', 0)
        ws['A10'] = "스마트스토어"
        ws['B10'] = report.get('sales', {}).get('naver', 0)
        ws['A11'] = "쿠팡"
        ws['B11'] = report.get('sales', {}).get('coupang', 0)
        
        # 배송 현황
        ws['A13'] = "🚚 배송 현황"
        ws['A13'].font = Font(bold=True, size=12)
        
        ws['A14'] = "발송 완료"
        ws['B14'] = report.get('shipping', {}).get('shipped', 0)
        ws['A15'] = "배송 완료"
        ws['B15'] = report.get('shipping', {}).get('delivered', 0)
        
        # 열 너비
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        if not filename:
            filename = f"daily_report_{report.get('date', datetime.now().strftime('%Y%m%d'))}.xlsx"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class CSVExporter:
    """CSV 내보내기"""
    
    @staticmethod
    def export_orders(orders: List[Dict], filename: str = None) -> ExportResult:
        """주문 목록 CSV 내보내기"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 헤더
        headers = ["번호", "채널", "주문번호", "주문자", "수령인", "연락처", "주소", "금액", "상태", "송장번호", "주문일시"]
        writer.writerow(headers)
        
        # 데이터
        for idx, order in enumerate(orders, 1):
            writer.writerow([
                idx,
                "스마트스토어" if order.get("channel") == "naver" else "쿠팡",
                order.get("channel_order_id", ""),
                order.get("buyer_name", ""),
                order.get("receiver_name", ""),
                order.get("receiver_phone", ""),
                order.get("receiver_address", ""),
                order.get("total_amount", 0),
                order.get("status", ""),
                order.get("tracking_number", ""),
                order.get("ordered_at", "")
            ])
        
        if not filename:
            filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue().encode('utf-8-sig'),  # BOM 포함
            content_type="text/csv; charset=utf-8"
        )
    
    @staticmethod
    def export_products(products: List[Dict], filename: str = None) -> ExportResult:
        """상품 목록 CSV 내보내기"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = ["SKU", "상품명", "재고", "임계값", "가격", "네이버ID", "쿠팡ID"]
        writer.writerow(headers)
        
        for product in products:
            writer.writerow([
                product.get("sku", ""),
                product.get("name", ""),
                product.get("stock_quantity", 0),
                product.get("stock_alert_threshold", 5),
                product.get("price", 0),
                product.get("naver_product_id", ""),
                product.get("coupang_product_id", "")
            ])
        
        if not filename:
            filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue().encode('utf-8-sig'),
            content_type="text/csv; charset=utf-8"
        )


class InvoicePDFGenerator:
    """송장 라벨 PDF 생성"""
    
    @staticmethod
    def generate_invoice_label(order: Dict, filename: str = None) -> ExportResult:
        """송장 라벨 PDF 생성"""
        if not PDF_AVAILABLE:
            return ExportResult(False, "", b"", "", "reportlab 라이브러리가 설치되지 않았습니다")
        
        output = io.BytesIO()
        
        # A6 사이즈 (105mm x 148mm) - 일반 송장 라벨 크기
        page_width = 105 * mm
        page_height = 148 * mm
        
        doc = SimpleDocTemplate(
            output,
            pagesize=(page_width, page_height),
            leftMargin=5*mm,
            rightMargin=5*mm,
            topMargin=5*mm,
            bottomMargin=5*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # 송장 정보 테이블
        data = [
            ["🏷️ 송장 라벨", ""],
            ["", ""],
            ["송장번호", order.get("tracking_number", "")],
            ["택배사", order.get("carrier", "CJ대한통운")],
            ["", ""],
            ["📦 수령인 정보", ""],
            ["이름", order.get("receiver_name", "")],
            ["연락처", order.get("receiver_phone", "")],
            ["주소", order.get("receiver_address", "")],
            ["우편번호", order.get("receiver_zipcode", "")],
            ["", ""],
            ["📋 주문 정보", ""],
            ["주문번호", order.get("channel_order_id", "")],
            ["채널", "스마트스토어" if order.get("channel") == "naver" else "쿠팡"],
            ["", ""],
            ["요청사항", order.get("buyer_memo", "-")],
        ]
        
        table = Table(data, colWidths=[30*mm, 60*mm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (0, 0), 12),
            ('FONTSIZE', (0, 5), (0, 5), 10),
            ('FONTSIZE', (0, 11), (0, 11), 10),
            ('FONTSIZE', (1, 2), (1, 2), 14),  # 송장번호 크게
            ('FONTNAME', (1, 2), (1, 2), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 2), (-1, 3), 0.5, colors.black),
            ('GRID', (0, 6), (-1, 9), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
            ('BACKGROUND', (0, 11), (-1, 11), colors.lightgrey),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        output.seek(0)
        
        if not filename:
            filename = f"invoice_{order.get('tracking_number', 'unknown')}.pdf"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=output.getvalue(),
            content_type="application/pdf"
        )


class BackupManager:
    """백업 관리"""
    
    @staticmethod
    def create_backup(data: Dict) -> ExportResult:
        """전체 데이터 백업 (JSON)"""
        backup_data = {
            "version": "1.0",
            "created_at": datetime.now().isoformat(),
            "data": data
        }
        
        content = json.dumps(backup_data, ensure_ascii=False, indent=2)
        filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        return ExportResult(
            success=True,
            filename=filename,
            content=content.encode('utf-8'),
            content_type="application/json"
        )
    
    @staticmethod
    def restore_backup(content: bytes) -> Dict:
        """백업 복원"""
        try:
            data = json.loads(content.decode('utf-8'))
            return {"success": True, "data": data.get("data", {})}
        except Exception as e:
            return {"success": False, "error": str(e)}
